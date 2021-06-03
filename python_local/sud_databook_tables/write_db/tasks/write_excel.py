import pandas as pd
import openpyxl as xl

def read_template_col(*, workbook, sheet_name, state_list, col = 'A', state_col_name = 'state', strip_chars=[]):
    """
    Function read_template_col to read the first (default) column of the given template to
    extract df with all state names in order as appear in excel, with corresponding row numbers
    (this allows for possible row breaks in template)
    
    params:
        workbook excel obj: workbook to read from
        sheet_name str: sheet to read
        state_list list: list of states to validate against
        col str: col to read, default is A (first col)
        state_col_name str: name of output df col with state names
        strip_chars list: optional list to provide any chars to strip from values
            e.g. ['*'] if there are asterisks next to names to indicate a set of states
    
    returns:
        df: df with list of states in order of template with state and rownum columns
    
    """
    
    workbook.active = workbook.sheetnames.index(sheet_name)
    sheet = workbook.active
    
    # add tuple for cell value and row to extracted list, remove all chars in strip_chars
    
    extracted = []
    for cell in sheet[col]:
        if cell.value:
            value = ''.join((filter(lambda x: x not in strip_chars, list(cell.value))))
            extracted.append((value, cell.row))
        
    # create df with state value and row number
    
    state_order = pd.DataFrame([value for value in extracted if value[0] in state_list],
                                columns = [state_col_name , 'rownum' ])
    
    return state_order

def write_cell(row_col, col_value, sheet, column, col_name):
    """
    Function write_cell to be applied row-wise to input df to write value to excel cell
    params:
        row_col str: name of df column that contains row # to write to
        col_value df value: value in df column to write to sheet
        sheet str: name of sheet
        column int: excel column # to write to
        col_name str: name of column writing to sheet

    If col_name ends in _stat, will apply comma with 1 decimal format, otherwise comma with 0 decimals

    returns:
        none

    """

    cellref = sheet.cell(row=row_col, column=column)
    cellref.value = col_value
    cellref.alignment = xl.styles.Alignment(horizontal='center')

    if col_name.endswith('_stat'):
        cellref.number_format = '###,##0.0'

    else:
        cellref.number_format = '###,##0'

def write_sheet(*, workbook, df, sheet_name, cols, scol, srow=None, row_col=None):
    """
        function write_text to write data to Excel sheet by cell
        
        params:
            workbook obj: open workbook object
            df df: dataframe to write
            sheet_name str: sheet name
            cols list: list of columns to write
            scol int: starting col
            srow int: optional starting row, if given will write sequentially at starting row, otherwise will use row_col given below
                must give either srow or row_col
            row_col str: optional value if specifying specific row numbers based on column with numbers in passed df 
                if not given, will write to each row in order beginning with srow
                if given, will use row number in that df col to determine where to write in excel sheet
        
        returns:
            None (writes to worksheet)
    
    """
    
    workbook.active = workbook.sheetnames.index(sheet_name)
    sheet = workbook.active

    # if srow is given, must create temp row_num col to pass to write_cell to determine row to write to

    if srow:
        row_col = 'row_num'
        df = df.copy().reset_index(drop=True)
        df[row_col] = df.index + srow

    # loop over all cols to write to table
    
    column=scol
    for col_name in cols:
        
        df[[row_col, col_name]].apply(lambda x: write_cell(*x, sheet = sheet, column = column, col_name = col_name), axis=1)

        column += 1